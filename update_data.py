# -*- coding: utf-8 -*-
"""
一鍵更新腳本
從本資料夾內的四個 xlsx 檔重新產生 data.js (4G) 與 data5g.js (5G)。
過濾條件：縣市=苗栗縣 且 鄉鎮=苗栗市
檔名會自動以 *LTE_CoBTS_CHT.xlsx / *LTE_CoCell_CHT.xlsx / *nrBts_DB_CHT.xlsx / *nrCell_DB_CHT.xlsx 匹配，
未來新 dump 直接放進來即可（支援不同日期前綴）。
"""
import openpyxl, json, os, glob, sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE = os.path.dirname(os.path.abspath(__file__))
# 過濾條件：以「基維股」欄位為準（苗栗基維股），不再限定單一鄉鎮
JIWEI = "苗栗基維股"

def find_file(pattern):
    files = glob.glob(os.path.join(BASE, pattern))
    if not files:
        raise FileNotFoundError(f"找不到 {pattern}，請確認檔案已放在本資料夾")
    return sorted(files)[-1]

def load_rows(path, sheet):
    wb = openpyxl.load_workbook(path, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise KeyError(f"{os.path.basename(path)} 內沒有 {sheet} 工作表")
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    hdr = [str(c).strip() if c is not None else "" for c in next(it)]
    rows = [dict(zip(hdr, r)) for r in it if any(c is not None for c in r)]
    wb.close()
    return hdr, rows

def load_rows_safe(path, sheet):
    """
    讀取 xlsx。若檔案被 Excel/OneDrive 鎖住（PermissionError），
    複製到暫存目錄再讀取。
    """
    try:
        return load_rows(path, sheet)
    except PermissionError:
        import tempfile, shutil
        tmp = os.path.join(tempfile.gettempdir(), "station_tmp",
                           os.path.basename(path))
        os.makedirs(os.path.dirname(tmp), exist_ok=True)
        shutil.copy2(path, tmp)
        _, rows = load_rows(tmp, sheet)
        return None, rows

def get(r, hdr, name):
    for i, h in enumerate(hdr):
        if h == name:
            v = r[h]
            return str(v).strip() if v is not None else ""
    return ""

def latlon_pairs(hdr, lat_base, lon_base, count=9):
    lat_cols = [lat_base] + [f"{lat_base}_{i}" for i in range(2, count + 1)]
    lon_cols = [lon_base] + [f"{lon_base}_{i}" for i in range(2, count + 1)]
    return lat_cols, lon_cols

def build_coords(row, hdr, lat_cols, lon_cols, n_stations):
    coords = []
    for i in range(n_stations):
        lat = get(row, hdr, lat_cols[i]) if i < len(lat_cols) else ""
        lon = get(row, hdr, lon_cols[i]) if i < len(lon_cols) else ""
        try:
            if lat and lon:
                coords.append({"lat": round(float(lat), 6), "lon": round(float(lon), 6)})
            else:
                coords.append(None)
        except (ValueError, TypeError):
            coords.append(None)
    return coords

def group_cells(cells, key_fn, field_map):
    """key_fn(cell_row) -> sector number; field_map -> output field names"""
    out = []
    for sec in sorted({key_fn(c) for c in cells if key_fn(c) is not None}):
        sec_cells = [field_map(c) for c in cells if key_fn(c) == sec]
        out.append({"sec": sec, "cells": sec_cells})
    return out


def station_town(cell_rows, hdr_cell, fallback):
    """站台鄉鎮：優先取該站台 cell 的鄉鎮（取多數），找不到就回傳 BTS 層級。"""
    towns = [str(get(c, hdr_cell, "鄉鎮")).strip() for c in cell_rows if get(c, hdr_cell, "鄉鎮")]
    towns = [t for t in towns if t]
    if towns:
        return max(set(towns), key=towns.count)
    return fallback or ""

def normalize(s):
    return s.replace(" ", "").replace("　", "").replace("-", "").replace("_", "")

import re as _re
def normalize_fold(s):
    """去空白/連字號，並將 'L<數字>' 與 'N<數字>' 視為相同（4G/5G 扇區字母不同）"""
    return _re.sub(r"[LN]\d{1,2}", "S", normalize(s))

def build_site_info(site_rows):
    """
    站台.xlsx：欄位 4G台號/5G台號/台名/地址（EAC 已改由 4 個 RMOD/SMOD 檔案提供）
    回傳 { "4g": { id: {站台名: {address} } }, "5g": { ... } }
    """
    info = {"4g": {}, "5g": {}}
    for r in site_rows:
        def g(k):
            v = r.get(k)
            return str(v).strip() if v is not None else ""
        g4 = g("4G台號")
        g5 = g("5G台號")
        tn = g("台名")
        addr = g("地址")
        if not tn:
            continue
        # 4G
        if g4:
            base4 = g4.rstrip("LUSl")
            bucket = info["4g"].setdefault(base4, {})
            entry = bucket.setdefault(tn, {"address": ""})
            if addr and not entry["address"]:
                entry["address"] = addr
        # 5G
        if g5:
            base5 = g5.rstrip("LUSl")
            bucket = info["5g"].setdefault(base5, {})
            entry = bucket.setdefault(tn, {"address": ""})
            if addr and not entry["address"]:
                entry["address"] = addr
    return info

def match_station(st, info_bucket):
    """在 info_bucket({站台名: entry}) 中模糊比對站台名 st，回傳 entry 或 None"""
    if st in info_bucket:
        return info_bucket[st]
    norm_st = normalize(st)
    fold_st = normalize_fold(st)
    for tn, entry in info_bucket.items():
        if normalize(tn) == norm_st:
            return entry
        if fold_st and (fold_st in normalize_fold(tn) or normalize_fold(tn) in fold_st):
            return entry
        if norm_st and (norm_st in normalize(tn) or normalize(tn) in norm_st):
            return entry
    return None

def parse_cell_desc(desc):
    """
    解析 RMOD 的 Description，格式如：
      Cell_013_Door / Cell_013_AC Power / Cell_013_ AC Power
    回傳 (cell_num, alarm)；若無 Cell_ 開頭回傳 (None, desc)
    """
    desc = (desc or "").strip()
    m = _re.match(r"[Cc]ell[_ ]*(\d+)[_ ]+(.*)$", desc)
    if m:
        return m.group(1).lstrip("0") or "0", m.group(2).strip()
    return None, desc

def build_eac_maps(cfg):
    """
    cfg: 指定 4 個 RMOD/SMOD 檔案的載入與解析
    每個檔案回傳 { id: { 名稱: {cell或"": [告警,...]} } }
    """
    maps = {}
    maps_used = {}
    for key, path, sheet, id_col, name_col, is_rmod in cfg:
        files = glob.glob(os.path.join(BASE, path))
        if not files:
            raise FileNotFoundError(f"找不到 {path}")
        fpath = sorted(files)[-1]
        hdr, rows = load_rows_safe(fpath, sheet)
        bucket = {}
        maps_used[key] = os.path.basename(fpath)
        for r in rows:
            def g(k):
                v = r.get(k)
                return str(v).strip() if v is not None else ""
            bid = g(id_col)
            name = g(name_col)
            desc = g("Description")
            lid = g("Cell_LID")
            if not bid or not name or not desc:
                continue
            entry = bucket.setdefault(bid, {}).setdefault(name, {})
            if lid:
                entry.setdefault("lids", []).append(lid)
            if is_rmod:
                cell, alarm = parse_cell_desc(desc)
                if cell is None:
                    continue  # RM 描述無 Cell 編號則略過
                d = entry.setdefault("rmod", {})
                d.setdefault(cell, []).append(alarm)
            else:
                d = entry.setdefault("smod", [])
                d.append(desc)
        maps[key] = bucket
    print("  EAC 使用檔案:", ", ".join(f"{k}={v}" for k, v in maps_used.items()))
    return maps

EAC_FILES = [
    # (key, glob_pattern, sheet, id_col, name_col, is_rmod)
    ("rmod4", "*LTE_RMOD_EAC_CHT.xlsx", "LTE_RMOD_EAC_CHT", "mrbtsId", "CellName_chs", True),
    ("smod4", "*LTE_SMOD_EAC_CHT.xlsx", "LTE_SMOD_EAC_CHT", "mrbtsId", "SiteName", False),
    ("rmod5", "*NR_RMOD_EAC_CHT.xlsx", "NR_RMOD_EAC_CHT", "mrbtsId", "SiteName", True),
    ("smod5", "*NR_SMOD_EAC_CHT.xlsx", "NR_SMOD_EAC_CHT", "mrbtsId", "SiteName", False),
]

def apply_eac(items, eac_maps, key_rmod, key_smod):
    """
    將 RMOD/SMOD 外部告警資訊合併到 items。
    輸出：
      it["eac"]     -> { 站台名: { "rmod": {cell_num: [告警...]} } }（RMOD，細胞階層）
      it["smod"]    -> { 站台名: [告警...] }（SMOD，站台階層）
      it["lid"]     -> { 站台名: [Cell_LID, ...] }（LID，站台階層）
    比對：以 id 精確 + 名稱模糊包含。
    """
    rmod = eac_maps.get(key_rmod, {})
    smod = eac_maps.get(key_smod, {})
    matched_cell = 0
    matched_smod = 0
    for it in items:
        bid = it["id"]
        # SMOD：以站台名(含 siteName/stations) 比對
        smod_bucket = smod.get(bid, {})
        if smod_bucket:
            # SMOD 為站台階層：任一欄位比對成功即套用到整個 item 的所有站
            site_hit = False
            hit_all = set()
            for name_key, listed in smod_bucket.items():
                target = match_any_name(name_key, it)
                if target is None:
                    continue
                if target == (it.get("siteName")):
                    site_hit = True
                else:
                    hit_all.add(target)
            site_alarms = set()
            site_lids = set()
            for name_key, listed in smod_bucket.items():
                site_alarms.update(listed.get("smod", []))
                site_lids.update(listed.get("lids", []))
            site_alarms = list(dict.fromkeys(site_alarms))
            site_lids = list(dict.fromkeys(site_lids))
            # SMOD 站台層級：套用至 item 的所有 station
            for st in (it.get("stations") or []):
                it.setdefault("smod", {})[st] = list(site_alarms)
                if site_lids:
                    it.setdefault("lid", {}).setdefault(st, []).extend(site_lids)
        # RMOD：以 CellName/SiteName 比對站台
        rmod_bucket = rmod.get(bid, {})
        for name_key, sub in rmod_bucket.items():
            target = match_any_name(name_key, it)
            if target is None:
                continue
            rmod_cells = sub.get("rmod", {})
            it.setdefault("eac", {}).setdefault(target, {})["rmod"] = {k: list(dict.fromkeys(v)) for k, v in rmod_cells.items()}
            lid_list = sub.get("lids", [])
            if lid_list:
                it.setdefault("lid", {}).setdefault(target, []).extend(lid_list)
            matched_cell += 1
        # 去重 LID
        if it.get("lid"):
            for st, lids in it["lid"].items():
                it["lid"][st] = list(dict.fromkeys(lids))
        if smod_bucket and it.get("stations"):
            matched_smod += 1
    return matched_cell, matched_smod

def match_any_name(name, it):
    """
    名稱 name 對 item 的站台做比對，回傳對應站台名或 None。
    優先匹配 stations 清單（cell 渲染用），其次 siteName。
    依序：精確、normalize 相等、normalize_fold 相等、雙向包含。
    """
    stations = it.get("stations") or []
    site = it.get("siteName")
    def _find(cands):
        if name in cands:
            return name
        norm = normalize(name)
        fold = normalize_fold(name)
        for c in cands:
            if normalize(c) == norm:
                return c
            if fold and fold == normalize_fold(c):
                return c
            if norm and (norm in normalize(c) or normalize(c) in norm):
                return c
        return None
    hit = _find(stations)
    if hit is not None:
        return hit
    if site:
        hit = _find([site])
        if hit is not None:
            return site
    return None

def enrich_items(items, site_info, id_key):
    """
    將站台地址資訊合併到 items（4G 或 5G）。（EAC 由 RMOD/SMOD 檔案另行處理）
    輸出：
      it["address"]  -> 第一個有地址的站之地址
      it["addr"]     -> { 站台名: 地址 }
    """
    matched = 0
    for it in items:
        bid = it["id"]
        info_bucket = site_info.get(id_key, {}).get(bid, {})
        if not info_bucket:
            continue
        addr_map = {}
        addr = ""
        for st in (it.get("stations") or []):
            entry = match_station(st, info_bucket)
            if entry is None:
                continue
            if entry.get("address"):
                addr_map[st] = entry["address"]
                if not addr:
                    addr = entry["address"]
        # 若站名對不上，但該 BTS 只有一筆資訊，則套用其地址(全站)
        if not addr and len(info_bucket) == 1:
            only = list(info_bucket.values())[0]
            if only.get("address"):
                addr = only["address"]
                for st in (it.get("stations") or []):
                    addr_map[st] = only["address"]
        if addr:
            it["address"] = addr
        if addr_map:
            it["addr"] = addr_map
            matched += 1
    return matched

def build_meter_map(meter_rows, id2x):
    """
    電號檔：欄位 電號▲/台號/台名（dict 形式，key 為欄位名）
    回傳 { lnBtsId: { 站台名: [電號, ...] } }
    比對順序：
      1. 台號去尾綴(L/U/S/l)後精確等於 lnBtsId → 命中該基地台
      2. 否則以台名對站台名做模糊(雙向包含)比對
    """
    ids4 = set(id2x.keys())
    meter_map = {}
    for r in meter_rows:
        def g(k):
            return str(r.get(k, "")).strip() if r.get(k) else ""
        meter_no = g("電號▲") or g("電號")
        th = g("台號")
        tn = g("台名")
        if not meter_no or not th:
            continue

        hit_id = None
        # 台號精確比對（去尾綴）
        if th in ids4:
            hit_id = th
        else:
            for suf in ["L", "U", "S", "l"]:
                base = th[:-1] if th.endswith(suf) else th
                if base in ids4:
                    hit_id = base
                    break

        station_hit = None
        if hit_id:
            x = id2x[hit_id]
            cands = x["stations"] or []
            # 台名精確匹配站台
            if tn:
                for st in cands:
                    if st == tn:
                        station_hit = st
                        break
                if station_hit is None:
                    for st in cands:
                        if normalize(st) == normalize(tn):
                            station_hit = st
                            break
                # 台名模糊匹配站台
                if station_hit is None:
                    for st in cands:
                        if tn and (normalize(tn) in normalize(st) or normalize(st) in normalize(tn)):
                            station_hit = st
                            break
            # 找不到具體站台 → 掛到第一個站台
            if station_hit is None:
                station_hit = cands[0] if cands else ""
        else:
            # 台名全域模糊比對
            for x in id2x.values():
                for st in x["stations"] or []:
                    if tn and (normalize(tn) in normalize(st) or normalize(st) in normalize(tn)):
                        hit_id = x["id"]
                        station_hit = st
                        break
                if hit_id:
                    break

        if hit_id and station_hit is not None:
            meter_map.setdefault(hit_id, {}).setdefault(station_hit, []).append(meter_no)
    return meter_map

# ============ 4G ============
print("== 4G ==")
co_bts = find_file("*LTE_CoBTS_CHT.xlsx")
co_cell = find_file("*LTE_CoCell_CHT.xlsx")
hdr_bts, bts_rows = load_rows(co_bts, "LTE_CoBTS_CHT")
hdr_cell, cell_rows = load_rows(co_cell, "LTE_CoCell_CHT")
print("  stations file:", os.path.basename(co_bts))
print("  cells file  :", os.path.basename(co_cell))

bts = [r for r in bts_rows if get(r, hdr_bts, "基維股") == JIWEI]
print(f"  過濾後站台數: {len(bts)}")

cell_map = {}
for r in cell_rows:
    lnbts = get(r, hdr_cell, "lnBtsId")
    st = get(r, hdr_cell, "CellName_CHS")
    if not lnbts or not st:
        continue
    cell_map.setdefault(lnbts, {}).setdefault(st, []).append(r)

# 電號檔（可選，找不到就跳過）
meter_rows = []
meter_files = glob.glob(os.path.join(BASE, "*電號*.xlsx"))
if meter_files:
    mf = sorted(meter_files)[-1]
    _, meter_rows = load_rows(mf, "工作表1")
    print(f"  電號檔: {os.path.basename(mf)} ({len(meter_rows)} 筆)")
else:
    print("  電號檔: 未找到，略過")

lat_cols, lon_cols = latlon_pairs(hdr_bts, "Lat", "Lon")
items4 = []
for r in bts:
    lnbts = get(r, hdr_bts, "lnBtsId")
    stations = get(r, hdr_bts, "lnBtsIdLL_NameChs").split("___") if get(r, hdr_bts, "lnBtsIdLL_NameChs") else []
    stations = [s.strip() for s in stations if s.strip()]
    secs = get(r, hdr_bts, "lnBtsIdLL_Sec").split("_") if get(r, hdr_bts, "lnBtsIdLL_Sec") else []
    coords = build_coords(r, hdr_bts, lat_cols, lon_cols, len(stations))
    cells_by_station = {}
    towns = []
    for s in stations:
        cell_rows_s = cell_map.get(lnbts, {}).get(s, [])
        towns.append(station_town(cell_rows_s, hdr_cell, get(r, hdr_bts, "鄉鎮")))
        def key_fn(c):
            try:
                return int(get(c, hdr_cell, "lnCelId")) // 10
            except (ValueError, TypeError):
                return None
        def field_map(c):
            obj = {
                "cel": get(c, hdr_cell, "lnCelId"),
                "cov": get(c, hdr_cell, "Coverage"),
                "rmod": get(c, hdr_cell, "rmod_cell"),
                "ant": get(c, hdr_cell, "AntType"),
                "az": get(c, hdr_cell, "Azimuth"),
                "mt": get(c, hdr_cell, "Mtilt"),
                "et": get(c, hdr_cell, "Etilt"),
            }
            try:
                lat, lon = float(get(c, hdr_cell, "Latitude")), float(get(c, hdr_cell, "Longitude"))
                obj["lat"] = round(lat, 6)
                obj["lon"] = round(lon, 6)
            except (ValueError, TypeError):
                pass
            return obj
        cells_by_station[s] = group_cells(cell_rows_s, key_fn, field_map)
    items4.append({
        "id": lnbts,
        "siteName": get(r, hdr_bts, "SiteName"),
        "siteNameCV": get(r, hdr_bts, "SiteNameCV"),
        "town": get(r, hdr_bts, "鄉鎮"),
        "towns": towns,
        "sec": get(r, hdr_bts, "lnBtsIdLL_Sec"),
        "secs": secs,
        "stations": stations,
        "coords": coords,
        "cells": cells_by_station,
        "nrBtsId": get(r, hdr_bts, "nrBtsId"),
        "ranType": get(r, hdr_bts, "RANtype"),
    })

# 電號比對（僅 4G）
if meter_rows:
    id2x = {x["id"]: x for x in items4}
    meter_map = build_meter_map(meter_rows, id2x)
    matched = 0
    for it in items4:
        m = meter_map.get(it["id"])
        if m:
            it["meters"] = m
            matched += 1
    print(f"  電號比對完成: {matched} 個站台有電號")

# 站台.xlsx（地址）
site_info = None
site_files = glob.glob(os.path.join(BASE, "站台.xlsx"))
if site_files:
    _, site_rows = load_rows_safe(site_files[0], "苗栗站台")
    site_info = build_site_info(site_rows)
    m4 = enrich_items(items4, site_info, "4g")
    print(f"  站台.xlsx 地址比對(4G): {m4} 個站台")

# RMOD/SMOD 外部告警
eac_maps = None
try:
    eac_maps = build_eac_maps(EAC_FILES)
    for k in ["rmod4", "smod4", "rmod5", "smod5"]:
        print(f"  EAC 檔 {k}: {len(eac_maps[k])} 個基地台")
except FileNotFoundError as e:
    print("  EAC 檔: 未完整找到，略過", e)

if eac_maps:
    c4, s4 = apply_eac(items4, eac_maps, "rmod4", "smod4")
    print(f"  RMOD 細胞標註(4G): {c4} 筆, SMOD 標註(4G): {s4} 個站台")

js4 = "const STATION_DATA = " + json.dumps(items4, ensure_ascii=False, indent=1) + ";\n"
with open(os.path.join(BASE, "data.js"), "w", encoding="utf-8") as f:
    f.write(js4)
print(f"  寫入 data.js: {len(items4)} 筆, {os.path.getsize(os.path.join(BASE,'data.js'))} bytes")

# ============ 5G ============
print("== 5G ==")
nr_bts = find_file("*nrBts_DB_CHT.xlsx")
nr_cell = find_file("*nrCell_DB_CHT.xlsx")
hdr5, bts5_rows = load_rows(nr_bts, "nrBts_DB_CHT")
hdr5c, cell5_rows = load_rows(nr_cell, "nrCell_DB_CHT")
print("  stations file:", os.path.basename(nr_bts))
print("  cells file  :", os.path.basename(nr_cell))

bts5 = [r for r in bts5_rows if get(r, hdr5, "基維股") == JIWEI]
print(f"  過濾後站台數: {len(bts5)}")

cell5_map = {}
for r in cell5_rows:
    nrbts = get(r, hdr5c, "nrBtsId")
    st = get(r, hdr5c, "nrCellName_chs")
    if not nrbts or not st:
        continue
    cell5_map.setdefault(nrbts, {}).setdefault(st, []).append(r)

lat5_cols, lon5_cols = latlon_pairs(hdr5, "Lat", "Lon")
items5 = []
for r in bts5:
    nrbts = get(r, hdr5, "nrBtsId")
    stations = get(r, hdr5, "nrBtsIdNN_NameChs").split("___") if get(r, hdr5, "nrBtsIdNN_NameChs") else []
    stations = [s.strip() for s in stations if s.strip()]
    secs = get(r, hdr5, "nrBtsIdNN_Sec").split("_") if get(r, hdr5, "nrBtsIdNN_Sec") else []
    coords = build_coords(r, hdr5, lat5_cols, lon5_cols, len(stations))
    cells_by_station = {}
    towns = []
    for s in stations:
        cell_rows_s = cell5_map.get(nrbts, {}).get(s, [])
        towns.append(station_town(cell_rows_s, hdr5c, get(r, hdr5, "鄉鎮")))
        def key_fn5(c):
            try:
                return int(get(c, hdr5c, "nrCellId")) // 100
            except (ValueError, TypeError):
                return None
        def field_map5(c):
            obj = {
                "cel": get(c, hdr5c, "nrCellId"),
                "rmod": get(c, hdr5c, "rMod"),
                "az": get(c, hdr5c, "Azimuth"),
                "mt": get(c, hdr5c, "M-tilt"),
                "et": get(c, hdr5c, "E-tilt"),
                "toff": get(c, hdr5c, "tiltOffset"),
                "cov": get(c, hdr5c, "Coverage"),
            }
            try:
                lat, lon = float(get(c, hdr5c, "Latitude")), float(get(c, hdr5c, "Longitude"))
                obj["lat"] = round(lat, 6)
                obj["lon"] = round(lon, 6)
            except (ValueError, TypeError):
                pass
            return obj
        cells_by_station[s] = group_cells(cell_rows_s, key_fn5, field_map5)
    items5.append({
        "id": nrbts,
        "siteName": get(r, hdr5, "SiteName"),
        "siteNameCV": get(r, hdr5, "SiteName"),
        "town": get(r, hdr5, "鄉鎮"),
        "towns": towns,
        "sec": get(r, hdr5, "nrBtsIdNN_Sec"),
        "secs": secs,
        "stations": stations,
        "coords": coords,
        "cells": cells_by_station,
        "ranType": get(r, hdr5, "RANtype"),
    })

if site_info:
    m5 = enrich_items(items5, site_info, "5g")
    print(f"  站台.xlsx 地址比對(5G): {m5} 個站台")

if eac_maps:
    c5, s5 = apply_eac(items5, eac_maps, "rmod5", "smod5")
    print(f"  RMOD 細胞標註(5G): {c5} 筆, SMOD 標註(5G): {s5} 個站台")

js5 = "const STATION_DATA_5G = " + json.dumps(items5, ensure_ascii=False, indent=1) + ";\n"
with open(os.path.join(BASE, "data5g.js"), "w", encoding="utf-8") as f:
    f.write(js5)
print(f"  寫入 data5g.js: {len(items5)} 筆, {os.path.getsize(os.path.join(BASE,'data5g.js'))} bytes")

print("\n完成！")
